#pip install xlrd


# Reading an excel file using Python

import openpyxl as opx

path = "Liste_des_livres_à_la_Bibliothèque.xlsx"

wb_obj = opx.load_workbook(path)

sheet_obj = wb_obj.active

print("Nombre de ligne ", sheet_obj.max_row)
dict_types = {}

for i in range(4, sheet_obj.max_row):
	cell_obj = sheet_obj.cell(row=i, column=8)
	key = cell_obj.value
	if key != None:
		key = key.strip()
		print("N° ", i, " ", key)
		if key in dict_types:
			dict_types[key] += 1
		else:
			dict_types[key] = 1
		

print(dict_types)

